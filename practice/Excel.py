#1500~1900
import openpyxl
i = 1900
try:
    wb = openpyxl.load_workbook('../RC/test_RC.xlsx')
    sheet = wb.active
    print("불러오기 완료.")
    data=[]
    data.append(i)
    for speed in range(1095,1905,5):
        if 1900>=speed>=1500:
            thruster_left = (1900-speed)/400*(i-1500)+speed
            thruster_right = speed - (i-1500)
            thruster_left_str = str(thruster_left)
            thruster_right_str = str(thruster_right)
            data.append(thrus
            ter_left_str + " " + thruster_right_str)
            data.pop(0)
            print(thruster_left,thruster_right)
            sheet.append(data)
            wb.save('test_RC.xlsx')

        elif 1500>speed>=1100:
            thruster_left = (speed-1100)/400*(i-1500)+speed
            thruster_right = (i-1700)*2/400*(1300-speed)+1300
            thruster_left_str = str(thruster_left)
            thruster_right_str = str(thruster_right)
            data.append(thruster_left_str + " " + thruster_right_str)
            data.pop(0)
            print(thruster_left,thruster_right)
            sheet.append(data)
            wb.save('test_RC.xlsx')
except:
    wb = openpyxl.Workbook('test_RC.xlsx')
    sheet = wb.active
    wb.save('test_RC.xlsx')
    print("새로운 파일을 만들었습니다.")

#1100~1500
import openpyxl
i = 1100
try:
    wb = openpyxl.load_workbook('../RC/test_RC.xlsx')
    sheet = wb.active
    print("불러오기 완료.")
    data=[]
    data.append(i)
    for speed in range(1095,1905,5):
        if 1900>=speed>=1500:
            thruster_left = speed - (1500 - i)
            thruster_right = (1900 - speed) / 400 * (1500 - i) + speed
            thruster_left_str = str(thruster_left)
            thruster_right_str = str(thruster_right)
            data.append(thruster_left_str + " " + thruster_right_str)
            data.pop(0)
            print(thruster_left,thruster_right)
            sheet.append(data)
            wb.save('test_RC.xlsx')

        elif 1500>speed>=1100:
            thruster_left = (1300 - i) * 2 / 400 * (1300 - speed) + 1300
            thruster_right = (speed - 1100) / 400 * (1500 - i) + speed
            thruster_left_str = str(thruster_left)
            thruster_right_str = str(thruster_right)
            data.append(thruster_left_str + " " + thruster_right_str)
            data.pop(0)
            print(thruster_left,thruster_right)
            sheet.append(data)
            wb.save('test_RC.xlsx')
except:
    wb = openpyxl.Workbook('test_RC.xlsx')
    sheet = wb.active
    wb.save('test_RC.xlsx')
    print("새로운 파일을 만들었습니다.")